import os
import pandas as pd
import openpyxl

import logging
from utils.logger import get_logger

formatter = logging.Formatter('%(message)s')

logger = get_logger('verify', formatter)

class account:
  def __init__(self, number, description):
    self.number = number
    self.description = description
    
    self.ingoing = []
    self.outgoing = []
  
  def total_in(self):
    return round(sum(self.ingoing), 2)
  
  def total_out(self):
    return round(sum(self.outgoing), 2)

  def balance(self):
    return round(self.total_in() - self.total_out(), 2)
  
  def update_target(self, amount):
    self.ingoing.append(amount)
  
  def update_origin(self, amount):
    self.outgoing.append(amount)

accounts = [
  account(8100, 'Umsatz'),
  account(703, 'Privater Ankauf'),
  account(1500, 'Helium Wallet HNT'),
  account(1100, 'Helium Wallet EUR'),
  account(4909, 'Fremdleistung'),
  account(1363, 'Transfer Helium Binance'),
  account(1501, 'Binance Wallet HNT'),
  account(1502, 'Binance Wallet USDT'),
  account(1101, 'Binance Wallet EUR'),
  account(1364, 'Transfer Binance Penta'),
  account(4970, 'Nebenkosten Geldtransfer'),
  account(1200, 'Penta'),
  account(2726, 'Ertraege Ein'),
  account(2725, 'Ertraege Aus'),
  account(2325, 'Verluste Ein'),
  account(2326, 'Verluste Aus')
]

export_list = [
  'datev_C&R_2022-1_Freber.xlsx',
  'datev_C&R_2022-2.xlsx',
  'datev_C&R_2022-3.xlsx',
  'datev_C&R_2022-4.xlsx',
]

export_df = pd.DataFrame()
for export_file in export_list:
  path = os.path.join('export', 'datev', export_file)
  part_df = pd.read_excel(path, header=None, skiprows=2, usecols=[0, 6, 7], names=['amount', 'target', 'origin'])

  export_df = export_df.append(part_df, ignore_index=True)



for index, event in export_df.iterrows():

  target_found = False
  origin_found = False

  if event.amount < 0:
    print(event)

  for account in accounts:
    if account.number == event.target:
      account.update_target(event.amount)
      target_found = True

    if account.number == event.origin:
      account.update_origin(event.amount)
      origin_found = True
  
  if not target_found or not origin_found:
    print(event)


logger.debug('-----------------------------------------')
for account in accounts:
  logger.debug('Account:')
  logger.debug(f'{account.number} - {account.description}')
  logger.debug(f'|{account.total_in()} || {account.total_out()}|')
  logger.debug(f'{account.balance()}')

# checks
mining_balance = round(accounts[0].total_out() + accounts[1].total_out() - accounts[2].total_in(), 2)
logger.debug(f'Mining 8100 + Transaktion 703 + Helium HNT 1500: {mining_balance}')

service_balance = round(accounts[3].total_in() - accounts[4].total_in(), 2)
logger.debug(f'Helium EUR 1100 - Fremdleistung 4909: {service_balance}')

neutral_service = round(accounts[12].total_in() + accounts[14].total_in() - accounts[6].total_out() -accounts[7].total_out(), 2)
logger.debug(f'2726 + 2325 - Binance HNT 1501 - Binance USDT 1502: {neutral_service}')

neutral_balance = round(accounts[2].total_out() - accounts[5].total_in() - neutral_service, 2)
logger.debug(f'1500 - 1363 - Neutral Service: {neutral_balance}')



file_path  = 'Verification.xlsx'

wb = openpyxl.load_workbook(file_path)
ws = wb.active

start_col = 1
for account in accounts:
  ws.cell(row=1, column=start_col, value=account.description)
  ws.cell(row=1, column=start_col+1, value=account.number)

  ws.cell(row=2, column=start_col, value='Summe Ein')
  ws.cell(row=2, column=start_col+1, value='Summer Aus')
  
  ws.cell(row=3, column=start_col, value=account.total_in())
  ws.cell(row=3, column=start_col+1, value=account.total_out())

  ws.cell(row=4, column=start_col, value='Ein')
  ws.cell(row=4, column=start_col+1, value='Aus')

  row_in = 5
  for entry in account.ingoing:
    ws.cell(row=row_in, column=start_col, value=entry)
    row_in += 1
  
  row_in = 5
  for entry in account.outgoing:
    ws.cell(row=row_in, column=start_col+1, value=entry)
    row_in += 1

  start_col += 3








wb.save(file_path)